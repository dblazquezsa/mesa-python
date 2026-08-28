import time
import pandas as pd
from modelo import Modelo_Desplazamiento
from openpyxl import Workbook, load_workbook
import os
import gc

# ------------------------------------------------------------
# Función auxiliar
# ------------------------------------------------------------

def agentes_insatisfechos(model):
    """Cuenta los agentes insatisfechos por tipo en el modelo."""
    insatisfechos = { -1:0, 0:0, 1:0 }
    for agent in model.agents:
        if agent.incomodidad(agent.pos, agent.tipo) > model.tolerance:
            insatisfechos[agent.tipo] += 1
    return insatisfechos

# ------------------------------------------------------------
# Montecarlo simple
# ------------------------------------------------------------

def run_montecarlo(n_runs=25, N1=500, N2=500, N3=500, tol=2.0, width=40, height=40): 
    resultados = []

    for run in range(1, n_runs+1):
        tiempo_inicio = time.time()
        model = Modelo_Desplazamiento(N1, N2, N3, width, height, tol, seed=run)

        # Ejecutar hasta equilibrio
        while model.steps_to_equilibrium < 20 and model.running and not model.in_equilibrium:
            model.step()

        # Medida de contacto final
        cm = model.contact_measure
        insatisfechos = agentes_insatisfechos(model)

        resultados.append({
            "Run": run,
            "Steps_to_Equilibrium": model.steps_to_equilibrium,
            # Contactos (6 medidas independientes)
            "C(-1,-1)": cm[-1][-1],
            "C(-1,0)": cm[-1][0],
            "C(-1,1)": cm[-1][1],
            "C(0,0)": cm[0][0],
            "C(0,1)": cm[0][1],
            "C(1,1)": cm[1][1],
            # Insatisfechos
            "Unsat_-1": insatisfechos[-1],
            "Unsat_0": insatisfechos[0],
            "Unsat_1": insatisfechos[1],
        })

        
        tiempo_fin = time.time()
        duracion = tiempo_fin - tiempo_inicio

        print(f"   → Experimento {run}/{n_runs} completado en {duracion:.2f} segundos.\n")

        model.running = False
        model.grid = None
        model.datacollector = None
        del model
        gc.collect()
    # Pasar a DataFrame y exportar
    
    return pd.DataFrame(resultados)
    #df.to_excel(salida, index=False)



# ------------------------------------------------------------
# Barrido sobre N2
# ------------------------------------------------------------


def barrido_N2(
    n_runs=20, N2_min=100, N2_max=150, paso=2, total=1500,
    tol=2.0, width=40, height=40, salida="barrido.xlsx"
):
    # Crear archivo si no existe
    if not os.path.exists(salida):
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Encabezados básicos
        columnas = ["N1", "N2", "N3","STE_mean","STE_std","CAA_mean","CAA_std","CAB_mean","CAB_std","CAC_mean","CAC_std", "CBB_mean","CBB_std","CBC_mean","CBC_std","CCC_mean","CCC_std","UnsatA_mean","UnsatA_std","UnsatB_mean","UnsatB_std","UnsatC_mean", "UnsatC_std"]
       
        ws.append(columnas)
        wb.save(salida)



    print(f"\n=== Ejecutando barrido desde N2={N2_min}, N2={N2_max} ===")

    # Abrir el libro una sola vez; se guarda tras cada iteración para no
    # perder progreso si el barrido se interrumpe, sin releerlo del disco.
    wb = load_workbook(salida)
    ws = wb.active

    for N2 in range(N2_min, N2_max + 1, paso):
        N1 = (total - N2) // 2
        N3 = N1

        print(f"\n=== Montecarlo para N2={N2} ===")

        df = run_montecarlo(n_runs, N1, N2, N3, tol, width, height)

        # crear fila
        fila = [N1, N2, N3]

        for col in df.columns:
            if col != "Run":
                fila.append(df[col].mean())
                fila.append(df[col].std())

        # Escribir fila en Excel
        ws.append(fila)
        wb.save(salida)

        # Limpiar memoria
        del df
        del fila

        gc.collect()

    wb.close()
    print(f"\n✔ Barrido completado. Guardado en {salida}")


